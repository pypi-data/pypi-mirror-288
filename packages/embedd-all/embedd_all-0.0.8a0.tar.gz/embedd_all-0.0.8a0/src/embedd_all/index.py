import openpyxl
import pandas as pd
from langchain.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter

def find_column_row(df):
    final_index = None
    final_row = []
    for index, row in df.iterrows():
        # Convert all items to strings for consistent processing
        str_row = row.apply(lambda x: str(x) if pd.notna(x) else '')
        non_nan_count = str_row.str.strip().astype(bool).sum()
        unnamed_count = sum(col.startswith('Unnamed') for col in str_row if isinstance(col, str))

        if non_nan_count > unnamed_count and non_nan_count > 1:
            final_index = index
            final_row = row[row.notna()].tolist()
        
        if non_nan_count > unnamed_count and non_nan_count > 2:
            return index, row[row.notna()].tolist()
        
    return final_index, final_row

def unmerge_and_populate(file_path, sheet_name):
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Create a copy of the merged cell ranges
    merged_cells_copy = list(sheet.merged_cells.ranges)

    # Dictionary to hold cell values to populate
    cell_values = {}

    # Iterate through the copied merged cell ranges
    for merged_cell in merged_cells_copy:
        min_col, min_row, max_col, max_row = merged_cell.min_col, merged_cell.min_row, merged_cell.max_col, merged_cell.max_row
        
        # Check if the merged cells are vertical (same column)
        if min_col == max_col:
            top_cell_value = sheet.cell(row=min_row, column=min_col).value

            # Unmerge the cells
            sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

            # Store the value to populate in unmerged cells
            for row in range(min_row, max_row + 1):
                cell_values[(row, min_col)] = top_cell_value

    # Populate the unmerged cells with the stored values
    for (row, col), value in cell_values.items():
        sheet.cell(row=row, column=col, value=value)

    # Save the workbook
    workbook.save(file_path)

    # Load the sheet into a pandas DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df

# Example usage
# file_path = '/Users/arnabbhattachargya/Desktop/data.xlsx'
# sheet_name = 'Sheet4'
# df = unmerge_and_populate(file_path, sheet_name)
# print(df)

def process_pdf(file_path):
    # create a loader
    loader = PyPDFLoader(file_path)
    # load your data
    data = loader.load()
    # Split your data up into smaller documents with Chunks
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=0)
    documents = text_splitter.split_documents(data)
    # Convert Document objects into strings
    texts = [str(doc) for doc in documents]
    return texts

def modify_excel_for_embedding(file_path, context):
    dfs = []

    xls = pd.ExcelFile(file_path)
    sheet_index = 0
    sheet_names = xls.sheet_names
    print("Sheet names:", sheet_names)
    for sheet in sheet_names:
        sheet_index = sheet_index + 1
        # df = pd.read_excel(file_path, sheet_name=sheet)
        df = unmerge_and_populate(file_path, sheet)

        # Find the row with column names
        column_row_index, column_names = find_column_row(df)
        if column_row_index is not None:
            # print(f"Column row found at index: {column_row_index}")
            print(f"Column names: {column_names}")
        else:
            print("No valid column row found.")

        # If column names are found, set the column names and drop the rows above it
        if column_row_index is not None:
            # Fill the beginning NaNs with placeholder column names
            column_names = ['Unnamed: ' + str(i) if pd.isna(col) else col for i, col in enumerate(df.iloc[column_row_index])]
            df.columns = column_names
            df = df.drop(range(column_row_index + 1)).reset_index(drop=True)
            
            # Drop columns starting with "Unnamed: "
            df = df.loc[:, ~df.columns.str.contains('Unnamed: ')]

        columns = column_names

        # Initialize the "summarized" column
        df["summarized"] = ""

        columns = df.columns

        # Iterate through each row to summarize the data
        for index, row in df.iterrows():
            summary = ""
            
            for col in columns:
                summary += col + ": " + str(row[col]).strip() + "; "
            
            if len(sheet_names) > 0:
                summary =  sheet + '/' + summary
            df.at[index, "summarized"] = context + '/' + summary

        # print("Processed Dataframe: ", df["summarized"])
        dfs.append(df["summarized"])
        print("Processed Sheet :", sheet_index)

    return dfs

