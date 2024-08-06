import os
import re
import csv
import io
from io import BytesIO
from openai import OpenAI
import fitz
import zipfile
import pandas as pd
from PIL import Image as PILImage
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
import numpy as np
from itertools import combinations
from scipy.spatial.distance import pdist, squareform
from google.cloud import storage
from google.auth.transport.requests import Request
from google.auth import impersonated_credentials
from google.oauth2 import service_account
from google.auth import default
import argparse

def initialize_storage_client():
    # Obtain default credentials (these are typically set via gcloud or environment variables)
    credentials, project = default()

    # Define the target service account to impersonate
    target_service_account = 'another-service-account@ai-artwork-parsing.iam.gserviceaccount.com'

    # Create impersonated credentials
    target_credentials = impersonated_credentials.Credentials(
        source_credentials=credentials,
        target_principal=target_service_account,
        target_scopes=['https://www.googleapis.com/auth/cloud-platform']
    )

    # Initialize the Google Cloud Storage client with impersonated credentials
    client = storage.Client(credentials=target_credentials)
    return client

def authenticate_implicit_with_adc(project_id):
    """
    When interacting with Google Cloud Client libraries, the library can auto-detect the
    credentials to use.

    // TODO(Developer):
    //  1. Before running this sample,
    //  set up ADC as described in https://cloud.google.com/docs/authentication/external/set-up-adc
    //  2. Replace the project variable.
    //  3. Make sure that the user account or service account that you are using
    //  has the required permissions. For this sample, you must have "storage.buckets.list".
    Args:
        project_id: The project id of your Google Cloud project.
    """

    # This snippet demonstrates how to list buckets.
    # *NOTE*: Replace the client created below with the client required for your application.
    # Note that the credentials are not specified when constructing the client.
    # Hence, the client library will look for credentials using ADC.
    storage_client = storage.Client(project=project_id)
    buckets = storage_client.list_buckets()
    print("Buckets:")
    for bucket in buckets:
        print(bucket.name)
    print("Listed all storage buckets.")

def get_file_format(file_path):
    format_signatures = {
        b'\x25\x50\x44\x46': 'PDF',  # %PDF
        b'\x50\x4B\x03\x04\x14\x00\x06\x00': 'Excel (XLSX)',
        b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1': 'Excel (XLS)',  # Older XLS files
        # Add more signatures as needed
    }

    max_signature_length = max(len(signature) for signature in format_signatures)

    with open(file_path, 'rb') as file:
        file_header = file.read(max_signature_length)

    for signature, file_format in format_signatures.items():
        if file_header.startswith(signature):
            return file_format

    # Additional check for XLSX (ZIP-based format)
    if zipfile.is_zipfile(file_path):
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            if '[Content_Types].xml' in zip_file.namelist():
                return 'Excel (XLSX)'

    return 'Unknown format'

def convert_pdf_to_text(pdf_path):
    doc = fitz.open(pdf_path)
    text = ""
    for page_num in range(doc.page_count):
        page = doc.load_page(page_num)
        text += page.get_text()
    return text

def convert_and_clean_excel_to_text(excel_path):
    # Read the Excel file
    excel_data = pd.read_excel(excel_path)

    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    cleaned_lines = []

    # Process each row in the Excel file
    for index, row in excel_data.iterrows():
        row_text_parts = []

        for col_index, cell_value in enumerate(row):
            # Check if the cell contains an image
            cell_coordinate = worksheet.cell(row=index + 2, column=col_index + 1).coordinate
            cell_has_image = any(image.anchor._from.col == col_index and image.anchor._from.row == index + 1 for image in worksheet._images)

            if cell_has_image:
                row_text_parts.append("[IMAGE HERE]")
            else:
                row_text_parts.append(str(cell_value))

        # Join the parts of the row text
        row_text = '\t'.join(row_text_parts)

        # Split row text by whitespace
        words = row_text.split()

        # Remove standalone 'nan' entries, but keep words that contain 'nan' as part of a word
        cleaned_words = [word for word in words if word != 'nan']

        # Join the cleaned words back into a string
        cleaned_line = ' '.join(cleaned_words)

        # Only add non-empty lines to the cleaned lines
        if cleaned_line:
            cleaned_lines.append(cleaned_line)

    # Join all cleaned lines into a single text
    cleaned_text = '\n'.join(cleaned_lines)

    return cleaned_text

def convert_to_text(INPUT, FORMAT):
    if FORMAT == 'PDF':
        text = convert_pdf_to_text(INPUT)
    elif FORMAT in ['Excel (XLSX)', 'Excel (XLS)']:
        text = convert_and_clean_excel_to_text(INPUT)
    print('File converted to text')
    return text

def extract_images_from_pdf(pdf_path):
    pdf_document = fitz.open(pdf_path)
    images_list = []

    for page_number in range(len(pdf_document)):
        page = pdf_document.load_page(page_number)
        images = page.get_images(full=True)

        # Extract image details and their positions
        image_positions = []
        for img in images:
            xref = img[0]
            base_image = pdf_document.extract_image(xref)
            image_bytes = base_image["image"]
            image = PILImage.open(BytesIO(image_bytes))
            bbox = page.get_image_bbox(img)
            image_positions.append((bbox, image))

        # Sort images by their y-coordinate to maintain order
        image_positions.sort(key=lambda x: x[0].y0)

        # Add sorted images to the main list
        for _, image in image_positions:
            images_list.append(image)

    return images_list

def extract_images_from_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    images_list = []

    # Handle the specified sheet
    sheet = wb.worksheets[0]
    for image in sheet._images:
        if isinstance(image, OpenPyxlImage):
            img_data = image._data()
            img_bytes = io.BytesIO(img_data)
            pil_img = PILImage.open(img_bytes)
            images_list.append((sheet.title, image.anchor._from.row, pil_img))  # Track sheet and position

    # Sort images by row (and possibly column) to maintain order
    images_list.sort(key=lambda x: x[1])
    sorted_images = [img for _, _, img in images_list]  # Extract sorted images

    return sorted_images

def find_most_similar_images(images, N):
    dimensions = [(image.size[0], image.size[1]) for image in images]

    # Calculate pairwise distances between all images
    distance_matrix = squareform(pdist(np.array(dimensions)))

    # Calculate the average distance from each image to all other images
    average_distances = np.mean(distance_matrix, axis=1)

    # Combine image indices with their average distances
    combined = list(zip(range(len(images)), average_distances))

    # Sort by average distance
    sorted_combined = sorted(combined, key=lambda x: x[1])

    # Select the indices of the top N images with the lowest average distances
    most_similar_indices = [idx for idx, _ in sorted_combined[:N]]
    most_similar_indices = sorted(most_similar_indices)

    # Retrieve the images corresponding to the most similar indices
    most_similar_images = [images[idx] for idx in most_similar_indices]

    return most_similar_images

def extract_images(INPUT, FORMAT, N):
    if FORMAT == 'PDF':
        image_list = extract_images_from_pdf(INPUT)
    if FORMAT in ['Excel (XLSX)', 'Excel (XLS)']:
        image_list = extract_images_from_excel(INPUT)
    image_list = find_most_similar_images(image_list, N)
    print('images extracted')
    return image_list

def upload_to_gcs(images, bucket):
    links = []
    bucket_name = 'art_images_eythos_595789'
    for i, image in enumerate(images):
        image_name = f'image_{i}.png'
        blob = bucket.blob(image_name)
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        buffer.seek(0)
        blob.upload_from_file(buffer, content_type='image/png')
        link = f'https://storage.googleapis.com/{bucket_name}/{image_name}'
        links.append(link)
    return links

def extract_info(client, input_text):
    # prompt to list out the basic information about each artwork
    list_artworks_prompt = f"""
    Based on the following information, list out the basic information about items, mostly artworks but ther can be exceptions, in a structured format as follows:
    Title:
    Artist:(if the item is not an artwork then say n/a)
    Height:(number only)
    Width:(number only, mark 0 if not provided)
    Depth:(number only, mark 0 if not provided)
    Unit:(the measurement unit for size)
    Material:
    Year:
    Value:(number only)
    Currency:(RMB/USD/EUR/KRW/etc.)
    Crated: (Y if crated/N if uncrated/(n/a))
    Framed: (Y if framed/N if unframed/(n/a))
    Owner:
    Do this for each item, output plain text, only the content I requested, some works might be identical, still treat them as different ones.
    If certain information is not provided output 'n/a'，only use 2 newlines to seperate them.
    {input_text}
    """

    # Make the second API call using the new interface
    list_artworks_response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "You are a helpful assistant who specializes in extracting information about items. You are good at keeping information clean and short."},
            {"role": "user", "content": list_artworks_prompt}
     ],
    )
    # print(list_artworks_response.choices[0].message.content)
    artworks_info = list_artworks_response.choices[0].message.content.strip()
    N = len(artworks_info.strip().split('\n\n'))
    print("Artwork information extracted")
    return artworks_info, N

def generate_csv_file(artworks_info, image_paths): # , image_paths, put this parameter in if image extraction needed
    # Define the input text
    input_text = artworks_info

    # Split the input text into individual artworks
    artworks = input_text.strip().split('\n\n')

    # extract images
    # Prepare the data for CSV writing
    csv_data = []
    for i, artwork in enumerate(artworks):
        # Use regular expressions to find each detail
        title = re.search(r'Title: (.+)', artwork).group(1).strip()
        artist = re.search(r'Artist: (.+)', artwork).group(1).strip()
        height = re.search(r'Height: (.+)', artwork).group(1).strip()
        width = re.search(r'Width: (.+)', artwork).group(1).strip()
        depth = re.search(r'Depth: (.+)', artwork).group(1).strip()
        unit = re.search(r'Unit: (.+)', artwork).group(1).strip()
        material = re.search(r'Material: (.+)', artwork).group(1).strip()
        year = re.search(r'Year: (.+)', artwork).group(1).strip()
        value = re.search(r'Value: (.+)', artwork).group(1).strip()
        currency = re.search(r'Currency: (.+)', artwork).group(1).strip()
        crated = re.search(r'Crated: (.+)', artwork).group(1).strip()
        framed = re.search(r'Framed: (.+)', artwork).group(1).strip()
        owner = re.search(r'Owner: (.+)', artwork).group(1).strip()

        # print(image_paths[i])
        # Possibly out of range
        # Append the row data to the csv_data list
        if i < len(image_paths):
             csv_data.append([
            title, image_paths[i], artist, height, width, depth, unit, material,
            year, value, currency, crated, framed, owner
        ]) 
        else:
            csv_data.append([
                title, "", artist, height, width, depth, unit, material,
                year, value, currency, crated, framed, owner
            ]) # image_paths[i], add this paramter when needed

    # Create an in-memory bytes buffer
    output = io.StringIO()

    # Write data to the in-memory bytes buffer as CSV
    writer = csv.writer(output)
    # Write the header
    writer.writerow([
        'Title', 'Image', 'Artist', 'Height', 'Width', 'Depth', 'Unit', 'Material',
        'Year', 'Value', 'Currency', 'Crated', 'Framed', 'Owner'
    ]) # 'Image', add this parameter when needed
    # Write the data rows
    writer.writerows(csv_data)

    # Move to the beginning of the BytesIO buffer
    output.seek(0)

    # Return the in-memory bytes buffer as a file-like object
    return output

def info_extraction(client, bucket, INPUT, PATH): # bucket, add when needed
  format = get_file_format(INPUT)
  print(f'{format} detected')
  text = convert_to_text(INPUT, format)
  artworks_info, N = extract_info(client, text)
  images = extract_images(INPUT, format, N)
  image_links = upload_to_gcs(images, bucket)
  csv = generate_csv_file(artworks_info, image_links) # , image_links add when needed
  df = pd.read_csv(csv)

  # TODO: Save the DataFrame to a CSV file in PATH
  df.to_csv(PATH, index=False)


# ai_client = OpenAI(api_key="sk-proj-BIs2MetN6kYzsavripaIT3BlbkFJH8UNk7dZoG1HVhSZMcUJ")

# INPUT = '/content/CIPL_ROPAC (AWJBLondon24)-fixed.xlsx'

# auth.authenticate_user()

# Set up Google Cloud Storage client
# gcs_client = storage.Client()
# bucket_name = 'art_images_eythos_595789'
# bucket = gcs_client.bucket(bucket_name)

# info_extraction(ai_client, INPUT) # bucket, add when needed

def main():
    parser = argparse.ArgumentParser(description='Extract artwork info and save to CSV.')
    parser.add_argument('input', type=str, help='Path to the input file')
    parser.add_argument('output', type=str, help='Path to save the output CSV file')
    args = parser.parse_args()
    
    ai_client = OpenAI(api_key="sk-proj-BIs2MetN6kYzsavripaIT3BlbkFJH8UNk7dZoG1HVhSZMcUJ")
    input_path = args.input
    output_path = args.output
    bucket_name = 'art_images_eythos_595789'
    authenticate_implicit_with_adc(project_id='ai-artwork-parsing')
    gcs_client = initialize_storage_client()
    bucket = gcs_client.bucket(bucket_name)
    
    output_csv_path = info_extraction(ai_client, bucket, input_path, output_path)
    print(f'The CSV file has been saved to {output_csv_path}')

if __name__ == '__main__':
    main()