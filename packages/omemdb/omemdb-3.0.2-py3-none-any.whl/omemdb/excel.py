from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from omemdb import Db
from omemdb.packages.omarsh import validate

dropdown_fill = PatternFill(start_color="CCFF99", fill_type="solid")


def generate_input_form(db: Db, path):
    # retrieve data
    db_d = dict()
    for table in db:
        table_d = dict()
        db_d[table.get_ref()] = table_d
        schema = table._dev_schema
        for k, v in schema.declared_fields.items():
            field_d = dict()
            table_d[k] = field_d
            choices = None
            if getattr(v, "validate", None) is not None and isinstance(v.validate, validate.OneOf):
                choices = v.validate.choices
            field_d["choices"] = choices

    wb = Workbook()
    wb.remove_sheet(wb.active)  # remove default sheet

    # choices worksheet
    ws = wb.create_sheet("choices")
    ws.cell(1, 1).value = "Table ref"
    ws.cell(1, 2).value = "Var name"
    ws.cell(1, 3).value = "Choices"

    # initialize and loop
    i = 1
    choices_d = dict()  # (table_ref, field_name): range
    for table_ref, table_data in db_d.items():
        for field_name, field_d in table_data.items():
            if field_d["choices"] is not None:
                i += 1
                ws.cell(i, 1).value = table_ref
                ws.cell(i, 2).value = field_name
                choices_nb = len(field_d["choices"])
                for k in range(choices_nb):
                    ws.cell(i, 3 + k).value = field_d["choices"][k]
                choices_d[(table_ref, field_name)] = f"choices!$C${i}:${get_column_letter(choices_nb + 3)}${i}"

    # table worksheets
    for table_ref, table_data in db_d.items():
        ws = wb.create_sheet(table_ref)
        for i, (field_name, field_data) in enumerate(table_data.items()):
            ws.cell(i + 1, 1).value = field_name
            if field_data["choices"] is not None:
                dv = DataValidation(type="list", formula1=choices_d[(table_ref, field_name)], allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(ws.cell(i + 1, 2))
                ws.cell(i + 1, 2).fill = dropdown_fill

    # save
    wb.save(path)
