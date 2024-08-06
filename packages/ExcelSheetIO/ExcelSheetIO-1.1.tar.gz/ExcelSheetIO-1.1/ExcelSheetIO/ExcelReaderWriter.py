'''
Developer :: soumyajitmahi7@gmail.com
'''

import os
import openpyxl
from robot.api import logger
import shutil  # for file operations
import random
from .CellStyler import CellStyler


class ExcelReaderWriter:
    '''
    The 'init' method reads the Excel sheet only once at the beginning of each test case execution. 
    This initial read operation loads the necessary data into memory, reducing the need for repeated I/O operations during the test run.
    '''
    def __init__(self, filename):
        self.__data = {}
        self.__originalFilePath = filename
        self.__wb = None
        try:
            wb = openpyxl.load_workbook(self.__originalFilePath, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = [cell.value for cell in sheet[1]]
                self.__data[sheet_name] = {}
                for row in sheet.iter_rows(min_row=2):
                    unique_identifier = row[0].value
                    if unique_identifier not in self.__data[sheet_name]:
                        self.__data[sheet_name][unique_identifier] = {}
                    for col_index, cell in enumerate(row):
                        col_name = headers[col_index]
                        self.__data[sheet_name][unique_identifier][col_name] = cell.value
            wb.close()
        except FileNotFoundError:
            logger.error(f"Error : File {self.__originalFilePath} not found")
        except PermissionError:
            logger.error(f"Error : Permission denied to access file {self.__originalFilePath}")
        except Exception as e:
            logger.error(f"Error : {e}")


    """
    This method retrieves data from a specific cell in an Excel sheet. 
    To use the 'readCellData' method, the following parameters are required:
        - Sheet Name: The name of the Excel sheet from which data is to be retrieved.
        - Unique Identifier: The specific unique identifier within the sheet for which data is needed.
        - Column Name: The column in the sheet that contains the desired data.
    By providing these parameters, the method can accurately locate and return the desired data from the Excel sheet.
    Return Type : String
    """
    def readCellData(self, sheetName, uniqueIdentifierInRow, columnHeaderName):
        """
        This method retrieves data from a specific cell in a given sheet.

        Parameters:
        sheetName (str): The name of the sheet from which to retrieve data.
        uniqueIdentifierInRow (str): The unique identifier used to locate the specific row in the sheet.
        columnHeaderName (str): The name of the column header used to locate the specific column in the sheet.

        Returns:
        var: The data from the specified cell if the sheet, row, and column exist.
        str: A message indicating that the column header, unique identifier, or sheet name was not found if either does not exist.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> cell_data = e.readCellData('Sheet1', 'E5', 'Job Title')
        """
        if sheetName in self.__data:
            if uniqueIdentifierInRow in self.__data[sheetName]:
                if columnHeaderName in self.__data[sheetName][uniqueIdentifierInRow]:
                    return self.__data[sheetName][uniqueIdentifierInRow][columnHeaderName]
                else:
                    return f"Column Header Name : '{columnHeaderName}' Not Found"
            else:
                return f"Unique Identifier : '{uniqueIdentifierInRow}' Not Found"
        else:
            return f"Sheet Name : '{sheetName}' Not Found"


    """
    This method is used to write data in excel sheet in the following steps:
    1. It writes data to a specific cell in an Excel sheet.
    2. Initially, it generates a copy of the original Excel sheet. This copy is stored in a folder named 'CopyFolder' and is named as '*_Temp*.xlsx'.
    3. The first write operation is executed on this temporary Excel file.
    4. After all write operations are completed, the data from the temporary file is transferred back into the original file.
    5. Finally, the temporary Excel file is removed after all operations are finished.
    This strategy ensures that the original Excel file remains unchanged until all write operations are successfully executed.
    Return Type : Void
    """
    def writeCellData(self, sheetName, uniqueIdentifierInRow, columnHeaderName , writable_data):
        """
        This method writes data to a specific cell in a given sheet of an Excel file.

        Parameters:
        sheetName (str): The name of the sheet where the data will be written.
        uniqueIdentifierInRow (str): The unique identifier used to locate the specific row in the sheet.
        columnHeaderName (str): The name of the column header used to locate the specific column in the sheet.
        writable_data: The data to be written to the specified cell.

        The method first creates a temporary copy of the original Excel file. It then loads the workbook and identifies the sheet to be modified. It locates the specific cell by matching the uniqueIdentifierInRow and columnHeaderName. Once the cell is identified, it writes the provided data to the cell. Finally, it saves and closes the workbook.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.writeCellData('Sheet1', 'E5', 'Job Title', 'Software Engineer')
        """
        self.__excelFilePath = self.__getTempFileName(self.__originalFilePath)   # temporary file
        self.__excelFilePath = self.__getTempFileName(os.path.join(self.__getCopyFolder_directory(), os.path.basename(self.__originalFilePath)))   # temporary file
        shutil.copyfile(self.__originalFilePath, self.__excelFilePath)  # create a copy
        self.__wb = openpyxl.load_workbook(self.__excelFilePath)
        sheet = self.__wb[sheetName]
        colIndex = 1
        maxRow = sheet.max_row
        # print(maxRow)
        for i in range(1, maxRow + 1):
            # print("Entered in Write For loop")
            testName = sheet.cell(i, 1).value
            # print(testName)
            if testName == uniqueIdentifierInRow:
                while (sheet.cell(row=1, column=colIndex).value != ''):
                    if (columnHeaderName == sheet.cell(row=1, column=colIndex).value):
                        break
                    colIndex = colIndex + 1
                sheet.cell(i, colIndex).value = writable_data
                # print(sheet.cell(i, colIndex).value)
                break
        # print("Come out of For Write loop")
        self.__wb.save(self.__excelFilePath)
        self.__wb.close()

        # Update the data attribute
        if sheetName in self.__data and uniqueIdentifierInRow in self.__data[sheetName]:
            self.__data[sheetName][uniqueIdentifierInRow][columnHeaderName] = writable_data
        self.__save_and_close(self.__excelFilePath)

    
    """
    This method generates a unique temporary filename.
    Here's how it works:
    1. It extracts the base name (the part of the filename without the extension) and the extension from the provided filename.
    2. It generates a random number between 1 and 1000.
    3. It constructs a new filename by appending '_temp', the random number, and the original extension to the base name.
    The resulting temporary filename maintains the original file type and is unlikely to conflict with existing files due to the inclusion of a random number.
    Return Type : String
    """
    def __getTempFileName(self, filename):
        base_name = os.path.splitext(filename)[0]  # get the base name of the file (without extension)
        extension = os.path.splitext(filename)[1]  # get the file extension
        random_number = random.randint(1, 1000)  # generate a random number between 1 and 1000
        temp_filename = f"{base_name}_temp{random_number}{extension}"
        return temp_filename
    

    """
    This method performs the following operations:
    1. It loads the workbook from the temporary Excel file.
    2. It saves any changes made to the workbook.
    3. It closes the workbook to free up system resources.
    4. It copies the data from the temporary file back to the original file. This ensures that the original file is updated only after all operations are successfully completed.
    5. It deletes the temporary file, as it is no longer needed.
    This approach helps maintain the integrity of the original file by preventing partial updates.
    Return Type : Void
    """
    def __save_and_close(self, excelFilePath):
        self.__wb = openpyxl.load_workbook(excelFilePath)
        self.__wb.save(excelFilePath)
        self.__wb.close()
        shutil.copyfile(excelFilePath, self.__originalFilePath)  # copy data back to the original file
        os.remove(excelFilePath)  # delete the temporary file


    """
    This method generates folder 'CopyFolder' in the following steps:
    1. It determines the current directory where this Python script is located.
    2. It constructs the path to a subdirectory named 'CopyFolder' within the current directory.
    3. If the 'CopyFolder' subdirectory does not exist, it creates it.
    4. It returns the path to the 'CopyFolder' subdirectory.
    This method is useful for managing temporary files that are created during the execution of this script.
    Return Type : String
    """
    def __getCopyFolder_directory(self):
        current_directory = os.path.dirname(os.path.realpath('__file__'))
        copyFolder_directory = os.path.join(current_directory, 'CopyFolder')
        if not os.path.isdir(copyFolder_directory):
            os.mkdir(copyFolder_directory)
        return copyFolder_directory


    """
    This method performs remove all temporary test data excel files in the following operations:
    1. It determines the directory where this Python script is located.
    2. It constructs the path to a subdirectory named 'CopyFolder' within the current directory.
    3. It iterates over all files in the 'CopyFolder' subdirectory.
    4. If a file name contains the string 'temp', it attempts to remove (delete) the file.
    5. If an error occurs during the file removal, it prints the error message.
    This method is useful for cleaning up temporary files that were created during the execution of this script.
    Return Type : Void
    """
    def removeTemporaryTestDataFiles(self):
        """
        This method removes all temporary test data files from a specified directory.

        It iterates over all files in the 'CopyFolder' directory located in the same directory as this script. If a file name contains the word 'temp', it attempts to remove that file.

        If an OSError occurs during the removal of a file, it logs the error using a logger.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.removeTemporaryTestDataFiles()
        """
        fileDir = os.path.dirname(os.path.realpath('__file__'))
        my_dir = os.path.join(fileDir, 'CopyFolder')
        if os.path.isdir(my_dir):
            for fname in os.listdir(my_dir):
                if 'temp' in fname:
                    try:
                        os.remove(os.path.join(my_dir, fname))
                    except OSError as error:
                        logger.error(error)


    """
    This method performs to read all data for a particular column in the following operations:
    1. It checks if the given sheet name exists in the data.
    2. If the sheet exists, it initializes an empty list to store the column data.
    3. It then iterates over all test cases in the sheet.
    4. If the given column name exists in the test case and its value is not None, it appends the value to the column data list.
    5. If the column data list is not empty after the iteration, it returns the list.
    6. If the column data list is empty, it returns a message indicating that no data was found in the given column in the given sheet.
    7. If the sheet does not exist in the data, it returns a message indicating that the sheet was not found.
    This method is useful for extracting all non-null data in a given column in a given sheet.
    Return Type : List
    """
    def readAllDataInGivenColumn(self, sheetName, columnHeaderName):
        """
        This method retrieves all data in a specified column of a given sheet.

        Parameters:
        sheetName (str): The name of the sheet from which to retrieve data.
        columnHeaderName (str): The name of the column header used to locate the specific column in the sheet.

        Returns:
        list: A list containing the data in the specified column if the sheet and column exist.
        str: A message indicating that no data was found in the specified column or the sheet name was not found if either does not exist.
        
        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> column_data = e.readAllDataInGivenColumn('Sheet1', 'Job Title')
        """
        if sheetName in self.__data:
            column_data = []
            for uniqueIdentifierInRow in self.__data[sheetName]:
                if columnHeaderName in self.__data[sheetName][uniqueIdentifierInRow] and self.__data[sheetName][uniqueIdentifierInRow][columnHeaderName] != None:
                    column_data.append(self.__data[sheetName][uniqueIdentifierInRow][columnHeaderName])
            if column_data:
                return column_data
            else:
                return f'No data found in column name-{columnHeaderName} in Sheetname-{sheetName}'
        else:
            return f'Sheet Name-{sheetName} not found in data'


    """
    This method fetch data for a particular row in the following operations:
    1. It checks if the given sheet name exists in the data.
    2. If the sheet exists, it checks if the given test case name exists in the sheet.
    3. If the test case exists, it retrieves all data for the test case. This data is a dictionary where the keys are column names and the values are cell values.
    4. It returns this dictionary, which represents all data in the given row.
    5. If the test case does not exist in the sheet, it returns a message indicating that the test case was not found in the given sheet.
    6. If the sheet does not exist in the data, it returns a message indicating that the sheet was not found.
    This method is useful for extracting all data in a given row in a given sheet.
    Return Type : Dictonary
    """
    def readAllDataInGivenRow(self, sheetName, uniqueIdentifierInRow):
        """
        This method retrieves all data in a specified row of a given sheet.

        Parameters:
        sheetName (str): The name of the sheet from which to retrieve data.
        uniqueIdentifierInRow (str): The unique identifier used to locate the specific row in the sheet.

        Returns:
        dict: A dictionary containing the data in the specified row if the sheet and row exist.
        str: A message indicating that the unique identifier or the sheet name was not found if either does not exist.
        
        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> row_data = e.readAllDataInGivenRow('Sheet1', 'E5')
        """
        if sheetName in self.__data:
            if uniqueIdentifierInRow in self.__data[sheetName]:
                row_data = self.__data[sheetName][uniqueIdentifierInRow]
                return row_data
            else:
                return f'Unique Identifier-{uniqueIdentifierInRow} not found in Sheet Name-{sheetName}'
        else:
            return f'Sheet Name-{sheetName} not found in data'


    '''
    This method is used to modify the color and font of a specific cell in an Excel sheet.
    Parameters:
        sheetName (str): The name of the sheet in the workbook.
        uniqueIdentifierInRow (str): The name of the test case, used to find the specific row in the sheet.
        columnHeaderName (str): The name of the column, used to find the specific cell in the row.
        cell_color (str, optional): The color to fill the cell with. Defaults to 'FFFFFF'.
        font_color (str, optional): The color of the font in the cell. Defaults to '000000'.
        font_type (bool, optional): If True, the font will be bold. Defaults to False.
    The method works by first creating a temporary copy of the original Excel file. It then loads the workbook from this temporary file and selects the specified sheet. It iterates over the rows in the sheet until it finds the row with the specified test case name. Then, it iterates over the columns in this row until it finds the specified column name. Once the specific cell is found, it changes the cell's fill color and font color as per the parameters. Finally, it saves and closes the workbook.
    This method is particularly useful for highlighting specific test cases in an Excel sheet, such as failed test cases in a test suite.
    Return Type : Void
    '''
    def modifyCellStyles(self, sheetName, uniqueIdentifierInRow, columnHeaderName):
        """
        This method creates a CellStyler object for a specific cell in a given sheet.

        Parameters:
        sheetName (str): The name of the sheet where the cell is located.
        uniqueIdentifierInRow (str): The unique identifier used to locate the specific row in the sheet.
        columnHeaderName (str): The name of the column header used to locate the specific column in the sheet.

        Returns:
        CellStyler: A CellStyler object initialized with the specified cell's details. This object can be used to modify the cell's styles.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> cell_styler = e.modifyCellStyles('Sheet1', 'E5', 'Job Title')
        """
        return CellStyler(self.__originalFilePath, self.__getTempFileName, self.__save_and_close, self.__getCopyFolder_directory, sheetName, uniqueIdentifierInRow, columnHeaderName)


    def resetCellStyles(self, sheetName, uniqueIdentifierInRow, columnHeaderName):
        """
        Resets the styles of a specific cell in an Excel sheet to default values.

        This method creates a CellStyler object for the specified cell, identified by its sheet name, unique identifier in the row, and column header name. It then sets the cell color to white, the font color to black, the font family to Calibri, the font style to normal, and the font size to 11. After applying these styles, it saves the changes to the Excel file.

        Parameters:
        sheetName (str): The name of the sheet where the cell is located.
        uniqueIdentifierInRow (str): The unique identifier in the row of the cell.
        columnHeaderName (str): The column header name of the cell.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.resetCellStyles('Sheet1', 'E7', 'Full Name')
        """
        cell_styler = CellStyler(self.__originalFilePath, self.__getTempFileName, self.__save_and_close, self.__getCopyFolder_directory, sheetName, uniqueIdentifierInRow, columnHeaderName)
        cell_styler.set_cell_color('white').set_font_color('black').set_font_family('Calibri').set_font_styles('normal').set_font_size(11).apply_styles().save_changes()