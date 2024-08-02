'''
Developer :: soumyajitmahi7@gmail.com
'''

import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Color
import shutil 
from .ColorPickers import *
from .ColorPickers import _get_hex_color

class CellStyler:
    def __init__(self, originalFilePath, getTempFileName,  save_and_close,  getCopyFolder_directory, sheetName, uniqueIdentifier, columnHeaderName):
        self.__originalFilePath = originalFilePath
        self.__getTempFileName = getTempFileName
        self.__save_and_close = save_and_close
        self.__getCopyFolder_directory = getCopyFolder_directory
        self.__sheetName = sheetName
        self.__uniqueIdentifier = uniqueIdentifier
        self.__columnHeaderName = columnHeaderName
        self.__styles = {}

    def set_cell_color(self, color):
        """
        Sets the cell color for the Excel cell.

        Parameters:
        color (str): The color to set for the cell.

        Returns:
        self: Returns the instance of the class for method chaining.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.modifyCellStyles().set_cell_color('red')
        """
        self.__styles['cell_color'] = color
        return self
    
    def set_font_family(self, font_family):
        """
        Sets the font family for the Excel cell.

        Parameters:
        font_family (str): The font family to set for the cell.

        Returns:
        self: Returns the instance of the class for method chaining.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.modifyCellStyles().set_font_family('Arial')
        """
        self.__styles['font_family'] = font_family
        return self

    def set_font_color(self, color):
        """
        Sets the font color for the Excel cell.

        Parameters:
        color (str): The color to set for the font.

        Returns:
        self: Returns the instance of the class for method chaining.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.modifyCellStyles().set_font_color('blue')
        """
        self.__styles['font_color'] = color
        return self

    def set_font_styles(self, *font_styles):
        """
        Sets the font styles for the Excel cell.

        Parameters:
        *font_styles (str): The font styles to set for the cell.

        Returns:
        self: Returns the instance of the class for method chaining.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.modifyCellStyles().set_font_styles('bold', 'italic', 'underline', 'double underline')
        """
        self.__styles['font_styles'] = font_styles
        return self

    def set_font_size(self, size):
        """
        Sets the font size for the Excel cell.

        Parameters:
        size (int): The font size to set for the cell.

        Returns:
        self: Returns the instance of the class for method chaining.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.modifyCellStyles().set_font_size(12)
        """
        self.__styles['font_size'] = size
        return self
    
    def apply_styles(self):
        """
        Applies the set styles to the Excel cell.

        Returns:
        _CellStyleSaver: Returns an instance of the _CellStyleSaver class which is used to save the changes to the Excel file.

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> cell_styler.set_cell_color('red').set_font_color('blue').set_font_type('Arial').set_font_size(12).apply_styles()
        """
        return _CellStyleSaver(self.__originalFilePath, self.__getTempFileName, self.__save_and_close, self.__getCopyFolder_directory, self.__sheetName, self.__uniqueIdentifier, self.__columnHeaderName, self.__styles)



class _CellStyleSaver:
    def __init__(self, originalFilePath, getTempFileName, save_and_close, getCopyFolder_directory, sheetName, uniqueIdentifier, columnHeaderName, styles):
        self.__originalFilePath = originalFilePath
        self.__getTempFileName = getTempFileName
        self.__save_and_close = save_and_close
        self.__getCopyFolder_directory = getCopyFolder_directory
        self.__sheetName = sheetName
        self.__uniqueIdentifier = uniqueIdentifier
        self.__columnHeaderName = columnHeaderName
        self.__styles = styles

    def save_changes(self):
        """
        This method saves changes made to an Excel file. It first creates a temporary copy of the original file, then opens the workbook and accesses the specified sheet. It iterates over the rows in the sheet until it finds the row with the unique identifier. Then, it iterates over the columns until it finds the column with the specified header name. Once the cell is identified, it applies the specified styles if any are provided. The styles can include cell color, font color, font type, and font size. Finally, it saves and closes the workbook.

        Parameters:
        None

        Returns:
        None

        Usage:
        >>> e = ExcelReaderWriter(...)
        >>> e.set_cell_color('red').set_font_color('blue').set_font_type('Arial').set_font_size(12).apply_styles().save_changes()
        """
        self.__excelFilePath = self.__getTempFileName(self.__originalFilePath)   # temporary file
        self.__excelFilePath = self.__getTempFileName(os.path.join(self.__getCopyFolder_directory(), os.path.basename(self.__originalFilePath)))   # temporary file
        shutil.copyfile(self.__originalFilePath, self.__excelFilePath)  # create a copy
        self.__wb = openpyxl.load_workbook(self.__excelFilePath)
        sheet = self.__wb[self.__sheetName]
        colIndex = 1
        maxRow = sheet.max_row

        for i in range(1, maxRow + 1):
            testName = sheet.cell(i, 1).value
            if testName == self.__uniqueIdentifier:
                while (sheet.cell(row=1, column=colIndex).value != ''):
                    if (self.__columnHeaderName == sheet.cell(row=1, column=colIndex).value):
                        break
                    colIndex = colIndex + 1

                cell = sheet.cell(i, colIndex)
                if self.__styles is not None:
                    font_color = _get_hex_color(self.__styles.get('font_color')) if 'font_color' in self.__styles else cell.font.color.rgb
                    bold = cell.font.b
                    italic = cell.font.i
                    underline = cell.font.u
                    font_size = self.__styles.get('font_size', cell.font.sz)
                    font_family = self.__styles.get('font_family', cell.font.name)

                    if 'font_styles' in self.__styles:
                        font_styles = self.__styles.get('font_styles', ('normal',))
                        for font_style in font_styles:
                            if font_style.lower() == 'bold':
                                bold = True
                            elif font_style.lower() == 'italic':
                                italic = True
                            elif font_style.lower() == 'underline':
                                underline = 'single'
                            elif font_style.lower() == 'double underline':
                                underline = 'double'
                            elif font_style.lower() == 'normal':
                                bold = False
                                italic = False
                                underline = None

                    cell.font = Font(name=font_family, color=Color(rgb=font_color), size=font_size, bold=bold, italic=italic, underline=underline)

                    if 'cell_color' in self.__styles:
                        cell_color = _get_hex_color(self.__styles.get('cell_color'))
                        cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")

                break

        self.__wb.save(self.__excelFilePath)
        self.__wb.close()
        self.__save_and_close(self.__excelFilePath)

